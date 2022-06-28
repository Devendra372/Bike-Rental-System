from kivy.lang import Builder
from kivymd.app import MDApp
from kivy.uix.screenmanager import ScreenManager, Screen
from reportlab.pdfgen import canvas
from kivymd.uix.button import MDFlatButton , MDRectangleFlatButton
from kivymd.uix.dialog import MDDialog
from kivymd.uix.label import MDLabel
from datetime import datetime , date
from kivy.core.window import Window
from kivy.lang import Builder
from openpyxl import Workbook, load_workbook

Window.size = (800, 800)




help_str = '''
#:import Factory kivy.factory.Factory
ScreenManager:
    BaseScreen:
    FinalPage:
    LoginScreen:
    
    
    

<BaseScreen>:
    name:'basescreen'
    MDScreen:
        md_bg_color: 239/255, 239/255, 239/255, 1
        name:"entrance"
    FitImage:
        source:"bike2.jpg"
        size_hint: (0.5,0.45)
    FitImage:
        source:"bike1.jpg"
        pos_hint:{"center_x":0.75}
        size_hint: (0.5,0.45)
        
   
    Image:
        source:'newlogo.png'
        pos_hint:{"center_x":0.5,"center_y": .9}
    MDLabel:
        text: "Welcome to SD Bike Rentals"
        font_size : 36
        pos_hint:{"center_x":0.5,"center_y": .7}
        # font_style:"H3"
        halign: "center"
        theme_text_color:"Custom"
        text_color:1, 0, 0, 
        opposite_colors: True


    MDRaisedButton:
        text:"Rent Bikes"
        elevation:10
        pos_hint:{"center_x": .5 , "center_y": 0.5}
        # size_hint: (0.5,0.1)
        size_hint:(0.15,0.01)
        bold:True
        on_press:
            root.manager.current = 'loginscreen'
            root.manager.transition.direction = 'left' 



<LoginScreen>:


    name:'loginscreen'
    MDLabel:
        text: "SD Bike Rentals"
        pos_hint:{"center_y": .9}
        font_style:"H3"
        halign: "center"
        theme_text_color:"Custom"
        text_color:0, 0, 0, 
    MDLabel:
        text: "Enter your Details"
        pos_hint:{"center_y": .8}
        # font_style:"H3"
        halign: "center"
        theme_text_color:"Custom"
        text_color:0, 0, 0, 

    MDTextField:
        id: name
        hint_text:"Enter your name"
        helper_text:'Required'
        helper_text_mode:  'on_error'
        pos_hint:{"center_x": .5 , "center_y": .7}
        current_hint_text_color:0, 0, 0, 1
        size_hint_x: .8
    MDTextField:
        id: phone_no
        hint_text:"Phone Number"
        helper_text:'Required'
        helper_text_mode:'on_error'
        pos_hint:{"center_x": .5 , "center_y": .6}
        current_hint_text_color:0, 0, 0, 1
        size_hint_x: .8
    MDTextField:
        id:no_bikes
        hint_text:"Number of Bikes"
        helper_text:'Required'
        helper_text_mode:  'on_error'
        pos_hint:{"center_x": .5 , "center_y": .5}
        current_hint_text_color:0, 0, 0, 1
        size_hint_x: .8

    MDTextField:
        id: duration
        hint_text:"On what basis do you want to rent bikes. (Hourly, Daily , Weekly)"
        helper_text:'Required'
        helper_text_mode:  'on_error'
        pos_hint:{"center_x": .5 , "center_y": .4}
        current_hint_text_color:0, 0, 0, 1
        size_hint_x: .8

    MDLabel:
        text: "We have currently "+str(app.displaystock())+" bikes available to rent."
        pos_hint:{"center_y": .3}
        # font_style:"H3"
        halign: "center"
        theme_text_color:"Custom"
        text_color:0, 0, 0, 

    
    MDRaisedButton:
        text:"Rent Bikes"
        pos_hint:{"center_x": .5 , "center_y": .2}
        size_hint_x: .5
        on_press:
            app.checkInput()

            # app.rentBikes()
            # app.popup()
       
            
                

    MDRectangleFlatButton:
        text: 'Go Back'
        pos_hint: {'center_x':0.85,'center_y':0.1}
        on_press:
            root.manager.current = 'basescreen'
            root.manager.transition.direction = 'right'



<FinalPage>:
    name:'final'
    
    MDLabel:
        text: " Thank You for choosing SD Bike Rentals "
        font_size : 36
        pos_hint:{"center_x":0.5,"center_y": .7}
        font_style:"H3"
        halign: "center"
        theme_text_color:"Custom"
        text_color:0, 0, 0, 

    MDLabel:
        text: " Your Order has been confirmed. You can receive your bikes from our nearest garage. "
        font_size : 20
        pos_hint:{"center_x":0.5,"center_y": .6}
        # font_style:"H3"
        halign: "center"
        theme_text_color:"Custom"
        text_color:0, 0, 0, 

    MDRaisedButton:
        text: 'Print Bill'
        pos_hint:{"center_x": .5 , "center_y": .3}
        size_hint:(0.15,0.01)
        on_press:
            app.generate_invoice()
    
        
    MDRaisedButton:
        text: 'Exit'
        size_hint:(0.15,0.01)
        pos_hint:{"center_x": .5 , "center_y": .2}
        # size_hint_x: .5
        # background_color =(1, 1, 1, 1)
        on_press:
            app.closeApp()


'''


class BaseScreen(Screen):
    pass

class LoginScreen(Screen):
    pass
    
class OnlyLoginScreen(Screen):
    pass


class FinalPage(Screen):
    pass

class ScreenManagement(ScreenManager):
    pass

        
        
sm = ScreenManager()
sm.add_widget(BaseScreen(name = 'basescreen'))
sm.add_widget(LoginScreen(name = 'loginscreen'))
sm.add_widget(OnlyLoginScreen(name='onlyloginscreen'))





class LoginApp(MDApp):
    
    def build(self):
        # self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette ='Red'
           
        self.strng = Builder.load_string(help_str)
 
     
        return self.strng
       

    def popup(self):
        self.dialog = MDDialog(
            text = "Do you want to Confirm ??",
            buttons = [
                MDFlatButton(text='CANCEL' , text_color =self.theme_cls.primary_dark, on_press= self.close),

                MDRectangleFlatButton( text='YES' ,text_color =self.theme_cls.primary_dark, on_press =self.rentBikes , on_release =self.close),
            ],
        )

        self.dialog.open()

    def popup2(self):
        self.dialog = MDDialog(
            text = "Please enter all details !!",
            buttons = [
                MDFlatButton(text='CANCEL' , text_color =self.theme_cls.primary_dark, on_press= self.close),

            ],
        )

        self.dialog.open()

    def popup3(self):
        self.dialog = MDDialog(
            text = "Please enter correct details !!",
            buttons = [
                MDFlatButton(text='CANCEL' , text_color =self.theme_cls.primary_dark, on_press= self.close),

            ],
        )

        self.dialog.open()


    def checkInput(self):
        name=self.strng.get_screen('loginscreen').ids.name.text
        phone = self.strng.get_screen('loginscreen').ids.phone_no.text

        duration = self.strng.get_screen('loginscreen').ids.duration.text
        number = self.strng.get_screen('loginscreen').ids.no_bikes.text

        if(len(number)==0 or len(duration)==0 or len(name)==0 or len(phone)==0):
            self.popup2()

        elif(number.isdigit()==False or phone.isdigit()==False):
            self.popup3()

        elif (int(number)>0 and duration=='Hourly' and int(number)<=self.stock):
            self.popup()
            
            # self.displayDetails()

        elif (int(number)>0 and duration=='Daily' and int(number)<=self.stock):
            self.popup()
            

        elif (int(number)>0 and duration=='Weekly' and int(number)<=self.stock):
            self.popup()

        else:
            self.popup3()
    

    
       

    def close(self,obj):
        self.dialog.dismiss()
    
    def closeApp(self):
        MDApp.get_running_app().stop()
        Window.close()


    

    def displaystock(self):
        """
        Displays the bikes currently available for rent in the shop.
        """
        wb = load_workbook('Database.xlsx')
        page = wb.active

        stock = page['G2']
        self.stock = stock.value
        # print("We have currently {} bikes available to rent.".format(self.stock))
        return self.stock


    def rentBikes(self,obj):

        duration = self.strng.get_screen('loginscreen').ids.duration.text
        number = self.strng.get_screen('loginscreen').ids.no_bikes.text


        if (int(number)>0 and duration=='Hourly' and int(number)<=self.stock):
            self.strng.get_screen('final').manager.current = 'final'
            self.strng.get_screen('final').manager.transition.direction = 'left'
            self.rentBikesOnHourlyBasis(int(number))
            self.updatecsv()
            # self.displayDetails()

        elif (int(number)>0 and duration=='Daily' and int(number)<=self.stock):
            self.strng.get_screen('final').manager.current = 'final'
            self.strng.get_screen('final').manager.transition.direction = 'left'
            self.rentBikeOnDailyBasis(int(number))
            self.updatecsv()



        elif (int(number)>0 and duration=='Weekly' and int(number)<=self.stock):
            self.strng.get_screen('final').manager.current = 'final'
            self.strng.get_screen('final').manager.transition.direction = 'left'
            self.rentBikeOnWeeklyBasis(int(number))
            self.updatecsv()



    def updatecsv(self):
        name = self.strng.get_screen('loginscreen').ids.name.text
        number = self.strng.get_screen('loginscreen').ids.phone_no.text
        bikes = self.strng.get_screen('loginscreen').ids.no_bikes.text
        basis = self.strng.get_screen('loginscreen').ids.duration.text
        now = datetime.now()

        data = [int(number) , name , int(bikes) , basis , now]

        wb = load_workbook('Database.xlsx')
        page = wb.active

        stock = page['G2']
        stock = stock.value - int(bikes)

        page['G2']=stock

        page.append(data)

        wb.save("Database.xlsx")
       
        
    def rentBikesOnHourlyBasis(self, n):
        """
        Rents a bike on hourly basis to a customer.
        """
        
        if n <= 0:
            print("Number of bikes should be positive!")
            return None
            
        
        elif n > self.stock:
            print("Sorry! We have currently {} bikes available to rent.".format(self.stock))
            return None
        
        else:
            now = datetime.now()                      
            print("You have rented a {} bike(s) on hourly basis today at {} hours.".format(n,now.hour))
            print("You will be charged $5 for each hour per bike.")
            print("We hope that you enjoy our service.")

            self.stock -= n
           
            return now
        

    def rentBikeOnDailyBasis(self, n):
        """
        Rents a bike on daily basis to a customer.
        """
        if n <= 0:
            print("Number of bikes should be positive!")
            return None

        elif n > self.stock:
            print("Sorry! We have currently {} bikes available to rent.".format(self.stock))
            return None
    
        else:
            now = datetime.now()                      
            print("You have rented {} bike(s) on daily basis today at {} hours.".format(n, now.hour))
            print("You will be charged $20 for each day per bike.")
            print("We hope that you enjoy our service.")

            self.stock -= n
            return now
        
    def rentBikeOnWeeklyBasis(self, n):
        """
        Rents a bike on weekly basis to a customer.
        """
        if n <= 0:
            print("Number of bikes should be positive!")
            return None

        elif n > self.stock:
            print("Sorry! We have currently {} bikes available to rent.".format(self.stock))
            return None        
        
        else:
            now = datetime.now()
            print("You have rented {} bike(s) on weekly basis today at {} hours.".format(n, now.hour))
            print("You will be charged $60 for each week per bike.")
            print("We hope that you enjoy our service.")
            self.stock -= n

            return now



    def generate_invoice(self):
        name = self.strng.get_screen('loginscreen').ids.name.text
        number = self.strng.get_screen('loginscreen').ids.phone_no.text
        bikes = self.strng.get_screen('loginscreen').ids.no_bikes.text
        basis = self.strng.get_screen('loginscreen').ids.duration.text
        if(basis=="Hourly"):
            price=20
        elif(basis=="Daily"):
            price=100
        else:
            price=500

        c = canvas.Canvas(str(name)+".pdf", pagesize=(200, 250), bottomup=0)
        c.setFillColorRGB(0, 0, 0)

        c.line(70, 22, 180, 22)
        c.line(5, 45, 195, 45)
        c.line(15, 120, 185, 120)
        c.line(35, 108, 35, 220)
        c.line(115, 108, 115, 220)
        c.line(135, 108, 135, 220)
        c.line(160, 108, 160, 220)
        c.line(15, 220, 185, 220)

        c.translate(10, 40)
        c.scale(1, -1)
        c.drawImage('newlogo.png', 0, 0, width=50, height=30)

        c.scale(1, -1)
        c.translate(-10, -40)

        c.setFont("Times-Bold", 10)
        c.drawCentredString(125, 20, 'SD Bike Rentals')

        c.setFont("Times-Bold", 5)
        c.drawCentredString(125, 30, 'COEP Shiivajinagar')
        c.drawCentredString(125, 35,  "Pune, India")
        c.setFont("Times-Bold", 6)
        c.drawCentredString(125, 42, "GST No: 10343293")

        c.setFont("Times-Bold", 8)
        c.drawCentredString(100, 55, "INVOICE")

        c.setFont("Times-Bold", 5)

        c.drawRightString(70, 70, "Invoice No. :")
        c.drawRightString(100, 70, "XXXX069")

        c.drawRightString(70, 80, "Date :")
        c.drawRightString(100, 80, str(date.today()))

        c.drawRightString(70, 90, "Customer Name :")
        c.drawRightString(113, 90, name)

        c.drawRightString(70, 100, "Phone No. :")
        c.drawRightString(100, 100, number)

        c.roundRect(15, 108, 170, 130, 10, stroke=1, fill=0)

        c.drawCentredString(25, 118, "S.No.")
        c.drawCentredString(75, 118, "Orders")
        c.drawCentredString(125, 118, "Price")
        c.drawCentredString(148, 118, "Qty.")
        c.drawCentredString(173, 118, "Total")

        c.drawCentredString(25, 133, '1')
        c.drawCentredString(75, 133, 'Bikes on '+str(basis)+' basis.')
        c.drawCentredString(125, 133, str(price))
        c.drawCentredString(148, 133, str(bikes))
        c.drawCentredString(173, 133, str(int(price)*int(bikes)))

        c.drawString(30, 230, "This is system generated invoice!!")
        c.drawString(30, 244, "*Please bring this bill when returning the bikes")


        # c.drawRightString(180, 228, 'Manager')
        c.drawRightString(180, 235, "Signature")

        c.showPage()
        c.save()
        
    
LoginApp().run()


